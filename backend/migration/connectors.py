"""
Data source connectors for legacy import.
Each connector returns: {"tables": [...], "columns": {table: [...]}, "rows": {table: [[...]]}}
"""
from __future__ import annotations
import csv
import json
import io
import sqlite3
from pathlib import Path
from typing import Any


# ── helpers ──────────────────────────────────────────────────────────────────

def _safe_rows(rows: list, limit: int = 2000) -> list:
    return rows[:limit]


def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


# ── CSV Connector ─────────────────────────────────────────────────────────────

def read_csv(file_bytes: bytes, filename: str = "data.csv") -> dict:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    columns = list(reader.fieldnames or (rows[0].keys() if rows else []))
    table = Path(filename).stem
    return {
        "tables": [table],
        "columns": {table: columns},
        "rows": {table: _safe_rows(rows)},
        "total": {table: len(rows)},
    }


# ── Excel Connector ───────────────────────────────────────────────────────────

def read_excel(file_bytes: bytes, filename: str = "data.xlsx") -> dict:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    tables: list[str] = []
    columns: dict = {}
    rows_map: dict = {}
    totals: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue
        header = [_str(c) if c is not None else f"col_{i}" for i, c in enumerate(all_rows[0])]
        data = []
        for row in all_rows[1:]:
            data.append({header[i]: _str(v) for i, v in enumerate(row) if i < len(header)})
        tables.append(sheet_name)
        columns[sheet_name] = header
        rows_map[sheet_name] = _safe_rows(data)
        totals[sheet_name] = len(data)

    return {"tables": tables, "columns": columns, "rows": rows_map, "total": totals}


# ── SQLite Connector ──────────────────────────────────────────────────────────

def read_sqlite(file_bytes: bytes) -> dict:
    import tempfile, os
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".sqlite3")
    try:
        tmp.write(file_bytes)
        tmp.close()
        conn = sqlite3.connect(tmp.name)
        conn.row_factory = sqlite3.Row
        tables = [r[0] for r in conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
        ).fetchall()]
        columns: dict = {}
        rows_map: dict = {}
        totals: dict = {}
        for t in tables:
            try:
                cur = conn.execute(f'SELECT * FROM "{t}" LIMIT 2000')
                cols = [d[0] for d in cur.description]
                data = [dict(zip(cols, r)) for r in cur.fetchall()]
                count = conn.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0]
                columns[t] = cols
                rows_map[t] = data
                totals[t] = count
            except Exception:
                pass
        conn.close()
        return {"tables": tables, "columns": columns, "rows": rows_map, "total": totals}
    finally:
        try:
            os.unlink(tmp.name)
        except Exception:
            pass


# ── JSON / MongoDB export Connector ──────────────────────────────────────────

def read_json(file_bytes: bytes, filename: str = "export.json") -> dict:
    data = json.loads(file_bytes.decode("utf-8", errors="replace"))
    table = Path(filename).stem

    if isinstance(data, list):
        rows = [r if isinstance(r, dict) else {"value": r} for r in data]
        cols = list(rows[0].keys()) if rows else []
        return {
            "tables": [table],
            "columns": {table: cols},
            "rows": {table: _safe_rows([{k: _str(v) for k, v in r.items()} for r in rows])},
            "total": {table: len(rows)},
        }

    if isinstance(data, dict):
        tables: list[str] = []
        columns: dict = {}
        rows_map: dict = {}
        totals: dict = {}
        for key, val in data.items():
            if isinstance(val, list) and val:
                rows = [r if isinstance(r, dict) else {"value": r} for r in val]
                cols = list(rows[0].keys()) if rows else []
                tables.append(key)
                columns[key] = cols
                rows_map[key] = _safe_rows([{k: _str(v) for k, v in r.items()} for r in rows])
                totals[key] = len(rows)
        if tables:
            return {"tables": tables, "columns": columns, "rows": rows_map, "total": totals}

    raise ValueError("Format JSON non reconnu (attendu: liste ou dict de listes)")


# ── External DB connectors (optional packages) ────────────────────────────────

def read_mysql(host: str, port: int, user: str, password: str, database: str,
               tables: list[str] | None = None) -> dict:
    try:
        import pymysql
    except ImportError:
        raise RuntimeError("pymysql non installé. Exécutez: pip install pymysql")
    conn = pymysql.connect(host=host, port=port, user=user, password=password,
                           database=database, charset="utf8mb4",
                           cursorclass=pymysql.cursors.DictCursor)
    return _read_sql_conn(conn, tables, quote_char="`")


def read_postgres(host: str, port: int, user: str, password: str, database: str,
                  tables: list[str] | None = None) -> dict:
    try:
        import psycopg2
        import psycopg2.extras
    except ImportError:
        raise RuntimeError("psycopg2 non installé. Exécutez: pip install psycopg2-binary")
    conn = psycopg2.connect(host=host, port=port, user=user, password=password,
                             dbname=database)
    return _read_sql_conn(conn, tables, quote_char='"', pg=True)


def read_sqlserver(host: str, port: int, user: str, password: str, database: str,
                   tables: list[str] | None = None) -> dict:
    try:
        import pyodbc
    except ImportError:
        raise RuntimeError("pyodbc non installé. Exécutez: pip install pyodbc")
    cs = f"DRIVER={{ODBC Driver 17 for SQL Server}};SERVER={host},{port};DATABASE={database};UID={user};PWD={password}"
    conn = pyodbc.connect(cs)
    return _read_sql_conn(conn, tables, quote_char='"')


def _read_sql_conn(conn: Any, tables: list[str] | None, quote_char: str = '"',
                   pg: bool = False) -> dict:
    cur = conn.cursor()
    q = quote_char
    # detect tables
    if pg:
        cur.execute("SELECT tablename FROM pg_catalog.pg_tables WHERE schemaname='public'")
        all_tables = [r[0] for r in cur.fetchall()]
    else:
        try:
            cur.execute("SHOW TABLES")
            all_tables = [list(r.values())[0] if hasattr(r, "values") else r[0] for r in cur.fetchall()]
        except Exception:
            cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
            all_tables = [r[0] for r in cur.fetchall()]

    selected = [t for t in all_tables if not tables or t in tables]
    columns_map: dict = {}
    rows_map: dict = {}
    totals: dict = {}

    for t in selected:
        try:
            cur.execute(f"SELECT * FROM {q}{t}{q} LIMIT 2000")
            cols = [d[0] for d in cur.description]
            if pg:
                rows = [dict(zip(cols, r)) for r in cur.fetchall()]
            else:
                rows = [dict(r) if hasattr(r, "keys") else dict(zip(cols, r)) for r in cur.fetchall()]
            count_cur = conn.cursor()
            count_cur.execute(f"SELECT COUNT(*) FROM {q}{t}{q}")
            total = count_cur.fetchone()[0]
            columns_map[t] = cols
            rows_map[t] = [{k: _str(v) for k, v in row.items()} for row in rows]
            totals[t] = total
        except Exception:
            pass

    conn.close()
    return {"tables": selected, "columns": columns_map, "rows": rows_map, "total": totals}


# ── Factory ───────────────────────────────────────────────────────────────────

class ConnectorFactory:
    @staticmethod
    def from_file(source_type: str, file_bytes: bytes, filename: str) -> dict:
        t = source_type.lower()
        if t == "csv":
            return read_csv(file_bytes, filename)
        if t in ("xlsx", "excel", "xls"):
            return read_excel(file_bytes, filename)
        if t in ("sqlite", "sqlite3", "db"):
            return read_sqlite(file_bytes)
        if t in ("json", "mongodb"):
            return read_json(file_bytes, filename)
        raise ValueError(f"Type de fichier non supporté: {source_type}")

    @staticmethod
    def from_db(source_type: str, host: str, port: int, user: str,
                password: str, database: str, tables: list[str] | None = None) -> dict:
        t = source_type.lower()
        if t == "mysql":
            return read_mysql(host, port, user, password, database, tables)
        if t in ("postgres", "postgresql"):
            return read_postgres(host, port, user, password, database, tables)
        if t in ("sqlserver", "mssql"):
            return read_sqlserver(host, port, user, password, database, tables)
        raise ValueError(f"Type de base non supporté: {source_type}")
